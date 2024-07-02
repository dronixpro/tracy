import os
import openpyxl
import requests
import re
import numpy as np
from sentence_transformers import SentenceTransformer
from functools import lru_cache
import streamlit as st


EXCEL_FILE_PATH = os.path.join(os.getcwd(), 'APP Layout.xlsx')
API_URL = "https://api-inference.huggingface.co/models/gpt2"
SENTENCE_TRANSFORMER_MODEL = 'all-MiniLM-L6-v2'

HUGGINGFACE_API_TOKEN = st.secrets["HUGGINGFACE_API_TOKEN"]

@lru_cache(maxsize=1)
def load_sentence_transformer():
    return SentenceTransformer(SENTENCE_TRANSFORMER_MODEL)

@lru_cache(maxsize=1)
def read_excel():
    if not os.path.exists(EXCEL_FILE_PATH):
        raise FileNotFoundError(f"Excel file not found at {EXCEL_FILE_PATH}")
    
    workbook = openpyxl.load_workbook(EXCEL_FILE_PATH, read_only=True, data_only=True)
    sheet = workbook.active
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    data = [[str(cell.value) if cell.value is not None else "" for cell in row] for row in sheet.iter_rows(min_row=2)]
    return headers, data

def create_vector_db(data, model):
    return [(" ".join(row), model.encode(" ".join(row))) for row in data]

def create_rag_prompt(query, vector_db, model, top_k=3):
    query_vector = model.encode(query)
    similarities = [np.dot(query_vector, entry[1]) / (np.linalg.norm(query_vector) * np.linalg.norm(entry[1])) for entry in vector_db]
    top_indices = np.argsort(similarities)[-top_k:][::-1]
    context = "\n".join([vector_db[i][0] for i in top_indices])
    
    return f"""Context information is below.
---------------------
{context}
---------------------
Given the resource information above and not prior knowledge, answer the question as a librarian. Provide a concise response without using markdown formatting or headers: {query}"""

def query_huggingface_api(prompt):
    headers = {"Authorization": f"Bearer {HUGGINGFACE_API_TOKEN}"}
    payload = {
        "inputs": prompt,
        "parameters": {"max_length": 150, "temperature": 0.7, "top_p": 0.9, "do_sample": True}
    }
    
    try:
        response = requests.post(API_URL, headers=headers, json=payload, timeout=10)
        response.raise_for_status()
        return response.json()
    except requests.RequestException as e:
        st.error(f"Error querying Hugging Face API: {e}")
        return None

def format_text(text):
    cut_off_phrase = "Given the resource information above and not prior knowledge"
    text = text.split(cut_off_phrase)[0].strip()
    
    def add_newline_after_link(match):
        return match.group(0) + "\n"
        
    url_pattern = r'https?://[^\s)"]+'
    
    formatted_text = re.sub(url_pattern, add_newline_after_link, text)
    
    formatted_text = re.sub(r'^#+\s*', '', formatted_text, flags=re.MULTILINE)
    formatted_text = re.sub(r'\*\*?(.*?)\*\*?', r'\1', formatted_text)
    formatted_text = re.sub(r'__?(.*?)__?', r'\1', formatted_text)
    
    return formatted_text.strip()

@st.cache_data
def main(query):
    model = load_sentence_transformer()
    headers, data = read_excel()
    vector_db = create_vector_db(data, model)
    rag_prompt = create_rag_prompt(query, vector_db, model)
    
    response = query_huggingface_api(rag_prompt)
    if response and isinstance(response, list) and len(response) > 0:
        text = response[0].get('generated_text', 'No response generated')
        formatted_response = format_text(text)
        return formatted_response
    else:
        return "I'm sorry, I couldn't generate a response. Please try rephrasing your question."

if __name__ == "__main__":
    st.title("Library Resource Assistant")
    user_query = st.text_input("What would you like to know?")
    if user_query:
        with st.spinner("Generating response..."):
            result = main(user_query)
        st.write(result)
